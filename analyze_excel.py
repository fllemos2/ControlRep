"""
Script para analisar a estrutura do arquivo Excel de Controle de Reprodução
"""
import openpyxl
import pandas as pd
import json
from pathlib import Path

# Caminho do arquivo Excel
excel_file = Path(r"C:\Users\fabiollemos\Downloads\Controle de Reprodução - Rev22 (1).xlsx")

print(f"Analisando: {excel_file}")
print(f"Arquivo existe: {excel_file.exists()}\n")

if not excel_file.exists():
    print("ERRO: Arquivo não encontrado!")
    exit(1)

# Carregar workbook
wb = openpyxl.load_workbook(excel_file)

# Listar todas as abas
print("=" * 80)
print("ABAS DO WORKBOOK:")
print("=" * 80)
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{i}. {sheet_name}")

print("\n" + "=" * 80)
print("DETALHES DE CADA ABA:")
print("=" * 80)

# Analisar cada aba
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n{'─' * 80}")
    print(f"ABA: {sheet_name}")
    print(f"{'─' * 80}")
    print(f"Dimensões: {ws.dimensions}")
    print(f"Linhas: {ws.max_row}, Colunas: {ws.max_column}")
    
    # Usar pandas para visualizar dados
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    print(f"\nColunas ({len(df.columns)}):")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i}. {col}")
    
    print(f"\nPrimeiras linhas:")
    print(df.head(10).to_string())
    
    print(f"\nDados vazios por coluna:")
    print(df.isnull().sum())

print("\n" + "=" * 80)
print("ANÁLISE COMPLETA")
print("=" * 80)
