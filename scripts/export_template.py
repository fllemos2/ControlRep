"""
Gera template Excel para carga de dados normalizados de Crias.
Exporta todos os registros atuais para revisão/normalização.

Uso:
    python scripts/export_template.py
Saída:
    data/template_crias.xlsx
"""

import os, sys

BACKEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend")
os.chdir(BACKEND_DIR)
sys.path.insert(0, BACKEND_DIR)

import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = os.path.join(os.path.dirname(__file__), "..", "data", "template_crias.xlsx")

COR_HEADER   = "2C5F2E"
COR_OBRIG    = "E8F5E9"
COR_VALIDADO = "E3F2FD"
COR_LIVRE    = "FFFFFF"
COR_FIXO     = "F5F5F5"
COR_VENDA    = "FFF8E1"   # amarelo claro — dados de venda

BORDA = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# (letra, título, largura, cor_fundo, formato_numero)
COLUNAS = [
    ("A", "ID",              8,   COR_FIXO,     None),
    ("B", "Reg. Nasc.",      14,  COR_OBRIG,    None),
    ("C", "Raca / Pelagem",  18,  COR_VALIDADO, None),
    ("D", "Sexo",            8,   COR_VALIDADO, None),
    ("E", "Data do Parto",   16,  COR_OBRIG,    "DD/MM/YYYY"),
    ("F", "Pai",             18,  COR_LIVRE,    None),
    ("G", "Status",          14,  COR_VALIDADO, None),
    ("H", "Nr Reg. Matriz",  16,  COR_FIXO,     None),
    ("I", "Vendido para",    22,  COR_VENDA,    None),
    ("J", "Data da Venda",   16,  COR_VENDA,    "DD/MM/YYYY"),
    ("K", "Peso (@)",        12,  COR_VENDA,    "0.00"),
    ("L", "Preco Real",      14,  COR_VENDA,    "#,##0.00"),
]

def make_header(ws):
    for col_letter, titulo, largura, _, _ in COLUNAS:
        cell = ws[f"{col_letter}1"]
        cell.value = titulo
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA
        ws.column_dimensions[col_letter].width = largura
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

def add_validations(ws, max_row):
    dv_raca = DataValidation(
        type="list", formula1='"Nel. Branca,Nel. Castanho,Nel. Pintado"',
        allow_blank=True, showDropDown=False, showErrorMessage=True,
        errorTitle="Valor invalido", error="Use: Nel. Branca, Nel. Castanho ou Nel. Pintado",
    )
    ws.add_data_validation(dv_raca)
    dv_raca.sqref = f"C2:C{max_row}"

    dv_sexo = DataValidation(
        type="list", formula1='"M,F"',
        allow_blank=True, showDropDown=False, showErrorMessage=True,
        errorTitle="Valor invalido", error="Use: M ou F",
    )
    ws.add_data_validation(dv_sexo)
    dv_sexo.sqref = f"D2:D{max_row}"

    dv_status = DataValidation(
        type="list", formula1='"No Pasto,Vendido,Morto,SUBMAT"',
        allow_blank=False, showDropDown=False, showErrorMessage=True,
        errorTitle="Valor invalido", error="Use: No Pasto, Vendido, Morto ou SUBMAT",
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = f"G2:G{max_row}"

def fill_data(ws, rows):
    for i, row in enumerate(rows, start=2):
        (id_, reg_nasc, raca, sexo, data_parto, pai,
         status, reg_matriz, vendido_para, data_venda, peso_venda, valor_venda) = row

        values = [id_, reg_nasc, raca, sexo, data_parto, pai,
                  status, reg_matriz, vendido_para, data_venda, peso_venda, valor_venda]

        for j, (col_letter, _, _, cor_fundo, fmt) in enumerate(COLUNAS):
            cell = ws[f"{col_letter}{i}"]
            cell.value = values[j]
            cell.border = BORDA
            cell.alignment = Alignment(vertical="center")
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            if fmt:
                cell.number_format = fmt
            if col_letter == "A":
                cell.font = Font(color="AAAAAA", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

def add_legenda(wb):
    ws = wb.create_sheet("Legenda")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    itens = [
        ("CAMPO",           "DESCRICAO",                                         True),
        ("ID",              "Identificador interno - NAO alterar",               False),
        ("Reg. Nasc.",      "Numero de registro de nascimento (unico)",           False),
        ("Raca / Pelagem",  "Nel. Branca | Nel. Castanho | Nel. Pintado",        False),
        ("Sexo",            "M = Macho  |  F = Femea",                          False),
        ("Data do Parto",   "Formato DD/MM/AAAA",                                False),
        ("Pai",             "Nome livre do reprodutor pai",                      False),
        ("Status",          "No Pasto | Vendido | Morto | SUBMAT",               False),
        ("Nr Reg. Matriz",  "Referencia da matriz mae - NAO alterar",            False),
        ("Vendido para",    "Nome do comprador (texto livre)",                   False),
        ("Data da Venda",   "Formato DD/MM/AAAA",                                False),
        ("Peso (@)",        "Peso em arrobas no momento da venda (ex: 12.5)",    False),
        ("Preco Real",      "Valor em reais recebido na venda (ex: 4500.00)",    False),
    ]
    for r, (campo, desc, header) in enumerate(itens, 1):
        a = ws.cell(r, 1, campo)
        b = ws.cell(r, 2, desc)
        if header:
            for c in (a, b):
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=COR_HEADER)
        a.border = b.border = BORDA

def main():
    conn = sqlite3.connect("cattle_control.db")
    db_rows = conn.execute("""
        SELECT
            c.id,
            c.numero_registro,
            c.raca_pelagem,
            c.sexo,
            c.data_nascimento,
            c.pai,
            c.status,
            m.numero_registro,
            c.vendido_para,
            c.data_venda,
            c.peso_venda,
            c.valor_venda
        FROM crias c
        JOIN matrizes m ON m.id = c.id_matriz
        ORDER BY c.data_nascimento, c.id
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crias"

    make_header(ws)
    fill_data(ws, db_rows)
    add_validations(ws, len(db_rows) + 10)
    add_legenda(wb)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"Template gerado: {os.path.abspath(OUTPUT)}")
    print(f"Total de linhas: {len(db_rows)}")

if __name__ == "__main__":
    main()
