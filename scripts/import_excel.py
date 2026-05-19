"""
Script de importação única do Excel legado para o banco SQLite.

Fases:
  1. Matrizes    (aba Matrizes, linha 5+)
  2. Reprodutores (pais únicos extraídos da aba Crias)
  3. Compradores  (compradores únicos extraídos da aba Crias)
  4. Crias        (aba Crias, linha 5+)

Uso:
  python scripts/import_excel.py --dry-run   # só exibe o que será importado
  python scripts/import_excel.py             # importação real
"""

import sys
import os
import argparse
from datetime import datetime, date

# Aponta para o backend como diretório de trabalho (SQLite usa path relativo)
BACKEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend")
os.chdir(BACKEND_DIR)
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from app.db.session import SessionLocal
from app.db.base import Base
from app.db.session import engine
from app.models.matriz import Matriz
from app.models.reprodutor import Reprodutor
from app.models.comprador import Comprador
from app.models.cria import Cria


EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "cattle_control.xlsx")


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def to_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def clean_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s and s.lower() not in ("none", "nan", "-") else None


def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


# ---------------------------------------------------------------------------
# Fase 1: Matrizes
# ---------------------------------------------------------------------------

def import_matrizes(db, dry_run):
    wb = load_workbook()
    ws = wb["Matrizes"]
    created = 0
    skipped = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        nr_raw = row[2]
        if nr_raw is None:
            continue

        numero_registro = str(int(nr_raw))
        total_crias = int(row[6]) if row[6] is not None else 0
        meses_por_cria = row[7]

        dados = {
            "numero_registro": numero_registro,
            "brinco": numero_registro,
            "raca": "Nelore",
            "status": "ativa",
            "total_crias": total_crias,
            "primeira_cria_data": to_date(row[3]),
            "ultima_cria_data": to_date(row[4]),
            "media_dias_intervalo": int(meses_por_cria * 30) if meses_por_cria else None,
            "observacoes": clean_str(row[8]),
        }

        if dry_run:
            print(f"  [matriz] {numero_registro}  crias={total_crias}")
            created += 1
            continue

        if db.query(Matriz).filter(Matriz.numero_registro == numero_registro).first():
            skipped += 1
            continue

        db.add(Matriz(**dados))
        created += 1

    if not dry_run:
        db.commit()

    print(f"Matrizes: {created} criadas, {skipped} já existiam")
    return created


# ---------------------------------------------------------------------------
# Fase 2: Reprodutores (pais únicos das Crias)
# ---------------------------------------------------------------------------

def import_reprodutores(db, dry_run):
    wb = load_workbook()
    ws = wb["Crias"]
    pais = {}

    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1] is None:
            continue
        pai = clean_str(row[7])
        if pai and pai not in pais:
            pais[pai] = pai

    created = 0
    skipped = 0

    for brinco in sorted(pais):
        if dry_run:
            print(f"  [reprodutor] {brinco}")
            created += 1
            continue

        if db.query(Reprodutor).filter(Reprodutor.brinco == brinco).first():
            skipped += 1
            continue

        db.add(Reprodutor(brinco=brinco, raca="Nelore"))
        created += 1

    if not dry_run:
        db.commit()

    print(f"Reprodutores: {created} criados, {skipped} já existiam")
    return created


# ---------------------------------------------------------------------------
# Fase 3: Compradores (únicos das Crias)
# ---------------------------------------------------------------------------

def import_compradores(db, dry_run):
    wb = load_workbook()
    ws = wb["Crias"]
    nomes = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1] is None:
            continue
        nome = clean_str(row[10])
        if nome:
            nomes.add(nome)

    created = 0
    skipped = 0

    for nome in sorted(nomes):
        if dry_run:
            print(f"  [comprador] {nome}")
            created += 1
            continue

        if db.query(Comprador).filter(Comprador.nome == nome).first():
            skipped += 1
            continue

        db.add(Comprador(nome=nome))
        created += 1

    if not dry_run:
        db.commit()

    print(f"Compradores: {created} criados, {skipped} já existiam")
    return created


# ---------------------------------------------------------------------------
# Fase 4: Crias
# ---------------------------------------------------------------------------

def import_crias(db, dry_run):
    wb = load_workbook()
    ws = wb["Crias"]

    # Índices lookup para resolver FKs
    matrizes_map = {}  # numero_registro → id
    reprodutores_map = {}  # brinco → id
    compradores_map = {}  # nome → id

    if not dry_run:
        for m in db.query(Matriz).all():
            matrizes_map[m.numero_registro] = m.id
        for r in db.query(Reprodutor).all():
            reprodutores_map[r.brinco] = r.id
        for c in db.query(Comprador).all():
            compradores_map[c.nome] = c.id

    created = 0
    skipped = 0
    sem_matriz = 0
    sem_data = 0
    registros_usados = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1] is None:
            continue

        brinco = clean_str(row[0])
        nr_matriz = str(int(row[1]))
        data_nasc = to_date(row[6])

        if data_nasc is None:
            sem_data += 1
            continue

        if dry_run:
            pai = clean_str(row[7])
            comp = clean_str(row[10])
            print(f"  [cria] brinco={brinco}  matriz={nr_matriz}  sexo={row[5]}  nasc={data_nasc}  pai={pai}  comprador={comp}")
            created += 1
            continue

        id_matriz = matrizes_map.get(nr_matriz)
        if id_matriz is None:
            sem_matriz += 1
            continue

        if brinco and db.query(Cria).filter(Cria.brinco == brinco).first():
            skipped += 1
            continue

        nr_reg_raw = row[3]
        try:
            numero_registro = str(int(nr_reg_raw)) if nr_reg_raw else None
        except (ValueError, TypeError):
            numero_registro = clean_str(nr_reg_raw)

        if numero_registro and numero_registro in registros_usados:
            numero_registro = None  # evita UNIQUE constraint
        elif numero_registro:
            registros_usados.add(numero_registro)

        pai_nome = clean_str(row[7])
        cria = Cria(
            brinco=brinco,
            id_matriz=id_matriz,
            id_reprodutor=reprodutores_map.get(pai_nome),
            pai=pai_nome,
            numero_ordem=int(row[2]) if row[2] else None,
            numero_registro=numero_registro,
            raca_pelagem=clean_str(row[4]),
            sexo=clean_str(row[5]),
            data_nascimento=data_nasc,
            origem=clean_str(row[9]) or "Cria da Fazenda",
            id_comprador=compradores_map.get(clean_str(row[10])),
            observacoes=clean_str(row[11]),
            data_venda=to_date(row[12]),
            peso_venda=to_float(row[13]),
            valor_venda=to_float(row[14]),
            indice_ganho=to_float(row[17]),
        )
        db.add(cria)
        created += 1

        if created % 100 == 0:
            db.flush()

    if not dry_run:
        db.commit()

    print(f"Crias: {created} criadas, {skipped} já existiam, {sem_matriz} sem matriz, {sem_data} sem data de nascimento")
    return created


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Importa dados do Excel para o banco")
    parser.add_argument("--dry-run", action="store_true", help="Apenas exibe o que seria importado, sem salvar")
    args = parser.parse_args()

    dry_run = args.dry_run

    print(f"\n{'=== DRY RUN ===' if dry_run else '=== IMPORTAÇÃO REAL ==='}")
    print(f"Arquivo: {os.path.abspath(EXCEL_PATH)}\n")

    if dry_run:
        db = None
    else:
        Base.metadata.create_all(bind=engine)
        db = SessionLocal()

    try:
        print("--- Fase 1: Matrizes ---")
        import_matrizes(db, dry_run)

        print("\n--- Fase 2: Reprodutores ---")
        import_reprodutores(db, dry_run)

        print("\n--- Fase 3: Compradores ---")
        import_compradores(db, dry_run)

        print("\n--- Fase 4: Crias ---")
        import_crias(db, dry_run)

        print("\nConcluído.")
    except Exception as e:
        if db:
            db.rollback()
        print(f"\nERRO: {e}")
        raise
    finally:
        if db:
            db.close()


if __name__ == "__main__":
    main()
